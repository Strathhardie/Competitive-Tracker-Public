import requests
import re
import csv
#pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from urllib.request import urlopen
from bs4 import BeautifulSoup
import time
#import pandas as pd


#The list of all potential FI's websites that we might visit
pages = ["https://www.cibc.com/en/special-offers/fall-savings-promotion.html",
"https://www.scotiabank.com/ca/en/personal/rates-prices/savings-account-rates.html",
# "https://www.scotiabank.com/ca/en/personal/bank-accounts/savings-accounts/momentum-plus-savings-account.html",
"https://www.tangerine.ca/en/landing-page/raptors"]


#This is write header to excel file
headers =['Bank','Account','Details','Special Offer']
book = Workbook()
sheet = book.active

for index, header in enumerate(headers):
    sheet.cell(row=1, column=index+1).value=header

#This is for reference to change background color code
#sheet['A1'].fill = PatternFill(start_color="FFC7CE", fill_type = "solid")

for index, link in enumerate(pages):


    #SCO_MOME
    if link == "https://www.scotiabank.com/ca/en/personal/rates-prices/savings-account-rates.html":
        sheet.cell(row=index+2, column=1).value="Scotiabank"
        page = requests.get(link)
        soup = BeautifulSoup(page.text, 'html.parser')

        #Account name
        account_name = soup.find(class_='heading').text.strip()
        sheet.cell(row=index+2, column=2).value=account_name

        # table = soup.find('div', class_ = '_momentum-plus-savings-rates')

        # #Offer detail
        # print(table.tbody)

        detail = soup.find('small').find('p').text.strip()
        # sheet.cell(row=index+2, column=3).value=detail

        perks = soup.find(class_ ='cmp cmp-text').find('p').text
        # sheet.cell(row=index+2, column=4).value=perks


    #CIBC
    elif link =="https://www.cibc.com/en/special-offers/fall-savings-promotion.html" :
        sheet.cell(row=index+2, column=1).value="CIBC"

        page = requests.get(link)
        soup = BeautifulSoup(page.text, 'html.parser')

        account_name= soup.find(text=re.compile('eAdvantage'))
        sheet.cell(row=index+2, column=2).value = account_name
        
        data = soup.find_all('div', class_ = 'longformtext base parbase')

        sheet.cell(row=index+2, column=3).value = data[4].text.strip()

        #Offer detail
        sheet.cell(row=index+2, column=4).value = data[3].text.strip()


    #Tangerine
    elif link =="https://www.tangerine.ca/en/landing-page/raptors" :
        sheet.cell(row=index+2, column=1).value="Tangerine"

        page= requests.get(link)
        soup = BeautifulSoup(page.text, 'html.parser')

        #Account name
        account_name= soup.find(text=re.compile('Savings'))
        sheet.cell(row=index+2, column=2).value=account_name

        data = soup.find_all('div', class_ = 'tngFlex')

        #Offer 1
        offer1 = soup.find(class_ = 'tngWrapper').find('p').text.strip()
        sheet.cell(row=index+2, column=3).value = offer1

        #Offer 2
        sheet.cell(row=index+2, column=4).value = data[2].text.strip()

book.save("specialOffer.xlsx")