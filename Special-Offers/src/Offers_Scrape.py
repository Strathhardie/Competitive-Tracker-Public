import requests
import re
import pandas as pd
import os
import xlsxwriter
import glob
import scraper
import json
import compare_Excel
import string
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from urllib.request import urlopen
from bs4 import BeautifulSoup
from datetime import date
from pathlib import Path



#This is write header to excel file
headers =['Bank','Account Name','Account Type', 'Monthly Fee', 'Special Offer','Expiry Date', 'Account Perks', 'Webiste']
book = Workbook()
sheet = book.active
sheet.title="Saving"
sheet2 = book.create_sheet("Chequing")

for index, header in enumerate(headers):
    sheet.cell(row=1, column=index+1).font = Font(bold=True)
    sheet.cell(row=1, column=index+1).value=header    
    sheet2.cell(row=1, column=index+1).font = Font(bold=True)
    sheet2.cell(row=1, column=index+1).value=header

# 1. Get dictionary return value from scraper.py for sepecial offer.
# print(json.dumps(scraper.get_special_offer_accounts(), indent=1))
special_offer_dict = scraper.get_special_offer_accounts()


def getList(dict): 
    return special_offer_dict.keys() 

rowNumSav = 2
rowNumChe = 2

def writeIntoWorkBook(SavOrCheSheet, bankName, rowIndex): 
    # SavOrCheSheet = sheet
    SavOrCheSheet.cell(row=rowIndex, column=1).alignment = Alignment(horizontal='center', vertical='center')
    SavOrCheSheet.cell(row=rowIndex, column=1).value=bankName

    SavOrCheSheet.cell(row=rowIndex, column=2).alignment = Alignment(horizontal='center', vertical='center')
    SavOrCheSheet.cell(row=rowIndex, column=2).value=special_offer_dict[x]['accounts'][y]['account_name'][0]  

    accountName = special_offer_dict[x]['accounts'][y]['account_name'][0]

    SavOrCheSheet.cell(row=rowIndex, column=3).alignment = Alignment(horizontal='center', vertical='center')
    SavOrCheSheet.cell(row=rowIndex, column=3).value=special_offer_dict[x]['accounts'][y]['account_category']

    SavOrCheSheet.cell(row=rowIndex, column=4).alignment= Alignment(wrapText=True)
    # SavOrCheSheet.cell(row=rowNum, column=4).alignment = Alignment(horizontal='left', vertical='center')
    writtenPara=""
    if not special_offer_dict[x]['accounts'][y]['fee']:
        writtenPara ="1. Monthly fee: $0"
    else:            
        for index, fee in enumerate(special_offer_dict[x]['accounts'][y]['fee']):           
            writtenPara += str(index+1)+". "+fee+"\n"   
    SavOrCheSheet.cell(row=rowIndex, column=4).value=writtenPara  

    SavOrCheSheet.cell(row=rowIndex, column=5).alignment= Alignment(wrapText=True)
    writtenPara=""
    if not special_offer_dict[x]['accounts'][y]['special_offer']:
        writtenPara ="No Data!"
    else:            
        for index, special_offer in enumerate(special_offer_dict[x]['accounts'][y]['special_offer']):           
            # writtenPara += str(index+1)+". "+special_offer+"\n"
            if accountName == 'The MomentumPLUS Savings Account':
                if index == 6:
                    break

            if accountName == 'RBC High Interest eSavings':
                s1 = str(special_offer)
                s2 = s1.replace('1:', ':')
                special_offer = s2
                if index == 1:
                    break
            if accountName == 'Savings Builder Account':
                if index == 2:
                    break
                s3 = str(special_offer)
                s4 = s3.replace("*16", ".")
                s5 = s4.replace("*17", "")
                s6 = s5.replace(". when", " when")
                special_offer = s6

            det1 = str(special_offer)
            cleaned = det1.replace('legal bug', '').rstrip(string.digits)
            # print("$$$"+cleaned)
            writtenPara += str(index + 1) + ". " + cleaned + "\n"

    SavOrCheSheet.cell(row=rowIndex, column=5).value=writtenPara     
    # SavOrCheSheet.cell(row=rowNum, column=4).value=special_offer_dict[x]['accounts'][y]['fee']

    writtenPara=""

    for index, detail in enumerate(special_offer_dict[x]['accounts'][y]['details']):
        if accountName == 'The MomentumPLUS Savings Account':
            st = str(detail)
            st1 = st.replace('account2.', 'account.')
            st2 = st1.replace('required4.', 'required')
            st3 = st2.replace('transfers6.', 'transfers')
            st4 = st3.replace('grow8.', 'grow')
            detail = st4
        det = str(detail)
        cleaned = det.replace('legal bug', '').rstrip(string.digits)
        writtenPara += str(index + 1) + ". " + cleaned + "\n"
    SavOrCheSheet.cell(row=rowIndex, column=7).value = writtenPara

    SavOrCheSheet.cell(row=rowIndex, column=8).alignment = Alignment(horizontal='center', vertical='center')
    SavOrCheSheet.cell(row=rowIndex, column=8).value=special_offer_dict[x]['accounts'][y]['account_url']


    rowIndex +=1





for x in getList(special_offer_dict):
    #Get Bank name
    bankName=special_offer_dict[x]['institution_name']

    for y in range(len(special_offer_dict[x]['accounts'])):
        #We need to seperate Saving account and Chequing account
        if special_offer_dict[x]['accounts'][y]['account_category']=="Saving Accounts":

            writeIntoWorkBook(sheet, bankName, rowNumSav)
            rowNumSav +=1

           
        else:


            writeIntoWorkBook(sheet2, bankName, rowNumChe)
            rowNumChe += 1



            


# #Today date in order to generate Excel timestamp
todayDate = str(date.today())





#Name convention for excel file "specialOfferYYYY-MM-DD"
book.save("specialOffer_"+todayDate+".xlsx")




#show different in excel compare last execute
compare_Excel.compare_changed_Special_Offer()