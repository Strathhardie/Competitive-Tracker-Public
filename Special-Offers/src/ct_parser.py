import requests
import re
import pandas as pd
import os
import xlsxwriter
import glob
import scraper
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from urllib.request import urlopen
from bs4 import BeautifulSoup
from datetime import date
import string
from pathlib import Path

headers = ['Bank', 'Account Name', 'Account Type', 'Monthly Fee', 'Special Offer', 'Expiry Date', 'Account Perks',
           'Website']
book = Workbook()
sheet = book.active

for index, header in enumerate(headers):
    sheet.cell(row=1, column=index + 1).font = Font(bold=True)
    sheet.cell(row=1, column=index + 1).value = header

special_offer_dict = scraper.get_special_offer_accounts()


def getList(dict):
    return special_offer_dict.keys()


rowNum = 2
for x in getList(special_offer_dict):
    # Get Bank name
    bankName = special_offer_dict[x]['institution_name']

    for y in range(len(special_offer_dict[x]['accounts'])):
        sheet.cell(row=rowNum, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=1).value = bankName

        sheet.cell(row=rowNum, column=2).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=2).value = special_offer_dict[x]['accounts'][y]['account_name'][0]

        accountName = special_offer_dict[x]['accounts'][y]['account_name'][0]

        sheet.cell(row=rowNum, column=3).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=3).value = special_offer_dict[x]['accounts'][y]['account_category']

        sheet.cell(row=rowNum, column=4).alignment = Alignment(wrapText=True)
        # sheet.cell(row=rowNum, column=4).alignment = Alignment(horizontal='left', vertical='center')
        writtenPara = ""
        if not special_offer_dict[x]['accounts'][y]['fee']:
            writtenPara = "1. Monthly fee: $0"
        else:
            for index, fee in enumerate(special_offer_dict[x]['accounts'][y]['fee']):
                writtenPara += str(index + 1) + ". " + fee + "\n"
        sheet.cell(row=rowNum, column=4).value = writtenPara

        sheet.cell(row=rowNum, column=5).alignment = Alignment(wrapText=True)
        writtenPara = ""
        if not special_offer_dict[x]['accounts'][y]['special_offer']:
            writtenPara = "No Data!"
        else:
            for index, special_offer in enumerate(special_offer_dict[x]['accounts'][y]['special_offer']):
                if accountName == 'The MomentumPLUS Savings Account':
                    if index == 6:
                        break

                if accountName == 'RBC High Interest eSavings':
                    s1 = str(special_offer)
                    s2 = s1.replace('1:', ':')
                    special_offer = s2
                    if index == 1:
                        break
                if accountName == 'Savings Builder Account':
                    if index == 2:
                        break
                    s3 = str(special_offer)
                    s4 = s3.replace("*16", ".")
                    s5 = s4.replace("*17", "")
                    s6 = s5.replace(". when", " when")
                    special_offer = s6

                det1 = str(special_offer)
                cleaned = det1.replace('legal bug', '').rstrip(string.digits)
                # print("$$$"+cleaned)
                writtenPara += str(index + 1) + ". " + cleaned + "\n"
        sheet.cell(row=rowNum, column=5).value = writtenPara
        # sheet.cell(row=rowNum, column=4).value=special_offer_dict[x]['accounts'][y]['fee']

        writtenPara = ""
        for index, detail in enumerate(special_offer_dict[x]['accounts'][y]['details']):
            if accountName == 'The MomentumPLUS Savings Account':
                st = str(detail)
                st1 = st.replace('account2.', 'account.')
                st2 = st1.replace('required4.', 'required')
                st3 = st2.replace('transfers6.', 'transfers')
                st4 = st3.replace('grow8.', 'grow')
                detail = st4
            det = str(detail)
            cleaned = det.replace('legal bug', '').rstrip(string.digits)
            writtenPara += str(index + 1) + ". " + cleaned + "\n"
        sheet.cell(row=rowNum, column=7).value = writtenPara

        sheet.cell(row=rowNum, column=8).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=8).value = r'=HYPERLINK("http://www.example.com","' + bankName + '")'
        print(bankName)

        rowNum += 1

############################################################
# The list of all potential FI's websites that we might visit
pages = ["https://www.cibc.com/en/special-offers/fall-savings-promotion.html",
         "https://www.scotiabank.com/ca/en/personal/rates-prices/savings-account-rates.html",
         # "https://www.scotiabank.com/ca/en/personal/bank-accounts/savings-accounts/momentum-plus-savings-account.html",
         "https://www.tangerine.ca/en/landing-page/raptors"]

# Today date in order to generate Excel timestamp
todayDate = str(date.today())

book.save("specialOffer" + todayDate + ".xlsx")
####################################################